import streamlit as st
import pandas as pd
import io
import chardet
from pathlib import Path

st.set_page_config(page_title="Excel Column Manager & Filter", layout="wide")
st.title("📊 Excel File Processor - Force Open & Clean")

def detect_encoding(raw_bytes):
    """Detect encoding for text files"""
    result = chardet.detect(raw_bytes)
    return result['encoding'] or 'utf-8'

def force_read_excel_or_anything(uploaded_file):
    """Try multiple methods to read any file into a DataFrame"""
    uploaded_file.seek(0)
    raw_bytes = uploaded_file.read(1024*1024)  # Read first 1MB for detection
    uploaded_file.seek(0)
    
    # Method 1: Try xlrd (old .xls) with corruption handling
    try:
        import xlrd
        book = xlrd.open_workbook(file_contents=uploaded_file.read(), ignore_workbook_corruption=True)
        uploaded_file.seek(0)
        sheet = book.sheet_by_index(0)
        data = []
        for row_idx in range(sheet.nrows):
            row = []
            for col_idx in range(sheet.ncols):
                cell = sheet.cell(row_idx, col_idx)
                row.append(cell.value)
            data.append(row)
        if data:
            df = pd.DataFrame(data[1:], columns=data[0] if data else None)
            return df, "xlrd (corruption ignored)"
    except Exception as e:
        st.warning(f"xlrd failed: {e}")
        uploaded_file.seek(0)
    
    # Method 2: Try openpyxl with read_only and data_only
    try:
        from openpyxl import load_workbook
        wb = load_workbook(uploaded_file, read_only=True, data_only=True, keep_links=False)
        sheet = wb.active
        rows = list(sheet.iter_rows(values_only=True))
        if rows:
            df = pd.DataFrame(rows[1:], columns=rows[0])
            return df, "openpyxl (read_only mode)"
    except Exception as e:
        st.warning(f"openpyxl failed: {e}")
        uploaded_file.seek(0)
    
    # Method 3: Try pandas with engine='xlrd' and ignore errors via catch
    try:
        df = pd.read_excel(uploaded_file, engine='xlrd', header=0)
        return df, "pandas + xlrd"
    except Exception as e:
        st.warning(f"pandas xlrd failed: {e}")
        uploaded_file.seek(0)
    
    # Method 4: Try pandas with engine='openpyxl'
    try:
        df = pd.read_excel(uploaded_file, engine='openpyxl', header=0)
        return df, "pandas + openpyxl"
    except Exception as e:
        st.warning(f"pandas openpyxl failed: {e}")
        uploaded_file.seek(0)
    
    # Method 5: Try reading as CSV (common mislabel)
    try:
        enc = detect_encoding(raw_bytes)
        df = pd.read_csv(uploaded_file, encoding=enc, on_bad_lines='skip')
        return df, f"CSV (encoding: {enc})"
    except Exception as e:
        st.warning(f"CSV failed: {e}")
        uploaded_file.seek(0)
    
    # Method 6: Try reading as HTML (extract first table)
    try:
        html_dfs = pd.read_html(uploaded_file)
        if html_dfs:
            return html_dfs[0], "HTML table extracted"
    except Exception as e:
        st.warning(f"HTML failed: {e}")
        uploaded_file.seek(0)
    
    # Method 7: Try raw text parsing (tab/space separated)
    try:
        enc = detect_encoding(raw_bytes)
        content = uploaded_file.read().decode(enc, errors='ignore')
        lines = content.splitlines()
        if lines:
            # Try splitting by tab or multiple spaces
            import re
            data = [re.split(r'\t| {2,}', line.strip()) for line in lines if line.strip()]
            if data:
                df = pd.DataFrame(data[1:], columns=data[0] if len(data) > 1 else None)
                return df, "Raw text (tab/space separated)"
    except Exception as e:
        st.warning(f"Raw text failed: {e}")
    
    raise ValueError("Could not read file with any method")

def process_dataframe(df, selected_columns, filters):
    """Apply column selection, reordering, and filters, skipping problematic columns"""
    # Ensure selected columns exist
    available_cols = [col for col in selected_columns if col in df.columns]
    missing = set(selected_columns) - set(available_cols)
    if missing:
        st.warning(f"Skipping missing columns: {missing}")
    
    if not available_cols:
        st.error("No valid columns selected.")
        return None
    
    # Select and reorder
    df = df[available_cols]
    
    # Apply filters
    if filters:
        for col, operator, value in filters:
            if col not in df.columns:
                st.warning(f"Filter column '{col}' not found, skipping filter")
                continue
            try:
                if operator == "Equals":
                    df = df[df[col].astype(str) == str(value)]
                elif operator == "Not Equals":
                    df = df[df[col].astype(str) != str(value)]
                elif operator == "Contains":
                    df = df[df[col].astype(str).str.contains(str(value), na=False)]
                elif operator == "Greater Than":
                    df = df[pd.to_numeric(df[col], errors='coerce') > float(value)]
                elif operator == "Less Than":
                    df = df[pd.to_numeric(df[col], errors='coerce') < float(value)]
                elif operator == "Is Empty":
                    df = df[df[col].isna() | (df[col].astype(str).str.strip() == '')]
                elif operator == "Not Empty":
                    df = df[df[col].notna() & (df[col].astype(str).str.strip() != '')]
            except Exception as e:
                st.warning(f"Error applying filter on '{col}': {e}. Skipping this filter.")
                continue
    return df

def main():
    st.markdown("""
    ### 💪 Force Open Mode
    This app will try **everything** to read your file:
    - Excel (xlrd, openpyxl) with corruption recovery
    - CSV (auto-detect encoding)
    - HTML tables
    - Raw text (tab/space separated)
    
    If a column causes problems, it will be skipped automatically.
    """)
    
    uploaded_file = st.file_uploader(
        "📂 Upload file (Excel, CSV, HTML, or text)",
        type=["xlsx", "xls", "csv", "html", "txt"],
        help="Large files supported. We'll force-open any format."
    )
    
    if uploaded_file is not None:
        # Get original size
        uploaded_file.seek(0, 2)
        original_size = uploaded_file.tell()
        uploaded_file.seek(0)
        st.info(f"📁 Original file size: {original_size / 1024:.2f} KB")
        
        # Force read
        with st.spinner("Attempting to read file using multiple methods (this may take a minute)..."):
            try:
                df, method = force_read_excel_or_anything(uploaded_file)
                st.success(f"✅ File successfully read using: **{method}**")
                st.write(f"Shape: {df.shape[0]} rows × {df.shape[1]} columns")
                
                # Show preview
                st.subheader("🔍 Data Preview (First 10 rows)")
                st.dataframe(df.head(10), use_container_width=True)
                
                # Column selection
                all_columns = df.columns.tolist()
                st.subheader("🎯 Column Management")
                st.markdown("Select columns in desired order. Uncheck to remove.")
                
                selected_columns = st.multiselect(
                    "Choose columns (order matters):",
                    options=all_columns,
                    default=all_columns[:min(10, len(all_columns))],  # Limit default to 10 columns to avoid overload
                    help="Click in sequence to define column order."
                )
                
                if not selected_columns:
                    st.warning("⚠️ Please select at least one column.")
                    return
                
                st.markdown(f"**Final column order:** {' → '.join(selected_columns)}")
                
                # Filter section
                st.subheader("🔍 Filter Rows (Optional)")
                with st.expander("Add filters"):
                    filters = []
                    num_filters = st.number_input("Number of filter conditions", min_value=0, max_value=10, value=0, step=1)
                    
                    for i in range(num_filters):
                        st.markdown(f"**Filter {i+1}**")
                        col1, col2, col3 = st.columns(3)
                        with col1:
                            filter_col = st.selectbox(f"Column", selected_columns, key=f"col_{i}")
                        with col2:
                            operator = st.selectbox(
                                f"Operator",
                                ["Equals", "Not Equals", "Contains", "Greater Than", "Less Than", "Is Empty", "Not Empty"],
                                key=f"op_{i}"
                            )
                        with col3:
                            if operator not in ["Is Empty", "Not Empty"]:
                                filter_value = st.text_input(f"Value", key=f"val_{i}")
                            else:
                                filter_value = None
                        
                        if operator not in ["Is Empty", "Not Empty"] and filter_value:
                            filters.append((filter_col, operator, filter_value))
                        elif operator in ["Is Empty", "Not Empty"]:
                            filters.append((filter_col, operator, None))
                
                # Process button
                if st.button("✅ Process & Download", type="primary"):
                    with st.spinner("Processing..."):
                        try:
                            processed_df = process_dataframe(df, selected_columns, filters)
                            if processed_df is None or processed_df.empty:
                                st.warning("No data remaining after processing.")
                                return
                            
                            # Save to Excel
                            output = io.BytesIO()
                            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                                processed_df.to_excel(writer, sheet_name="Processed", index=False)
                            
                            processed_bytes = output.getvalue()
                            new_size = len(processed_bytes)
                            reduction = (1 - new_size / original_size) * 100 if original_size > 0 else 0
                            
                            col1, col2, col3 = st.columns(3)
                            col1.metric("Original Size", f"{original_size / 1024:.2f} KB")
                            col2.metric("New Size", f"{new_size / 1024:.2f} KB")
                            col3.metric("Size Reduction", f"{reduction:.1f}%")
                            
                            st.subheader("📋 Processed Data Preview (First 5 rows)")
                            st.dataframe(processed_df.head(5), use_container_width=True)
                            
                            st.download_button(
                                label="📥 Download Cleaned Excel File",
                                data=processed_bytes,
                                file_name=f"cleaned_{uploaded_file.name}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                            )
                            
                            st.success(f"✅ Success! Final rows: {len(processed_df)}")
                            
                        except Exception as e:
                            st.error(f"Processing error: {e}")
                            st.info("Try selecting fewer columns or removing problematic columns.")
                            
            except Exception as e:
                st.error(f"❌ Cannot read file: {e}")
                st.info("""
                The file appears to be severely corrupted or not a tabular format.
                Suggestions:
                - Try opening the file in Excel/LibreOffice and re-saving as .xlsx
                - If it's a text file, ensure it has consistent delimiters (comma, tab, etc.)
                - Upload a different file
                """)

if __name__ == "__main__":
    main()